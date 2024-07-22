from typing import List

from files.models import File
from openpyxl import load_workbook


class ResultsProcessor:

    def __init__(self):
        self.results = None
        pass

    @staticmethod
    def fetch_results(num_results: int) -> List[File]:
        query_set = File.objects.filter(
            file__contains="BrightonPhoenix",
            file__iendswith=".xlsx").order_by("-created_date")[:num_results]

        return list(query_set)

    def process(self, curr_file: File, prev_file: File):
        curr_workbook = load_workbook(filename=curr_file.file.path)
        prev_workbook = load_workbook(filename=prev_file.file.path)

        curr_results = curr_workbook["Results (30 days)"]
        prev_results = prev_workbook["Results (30 days)"]

        sheet = curr_workbook.create_sheet(title="NEW RESULTS", index=0)
        sheet = self._append_headers(sheet)
        self._append_new_rows(curr_results, prev_results, sheet)

        self.results = curr_workbook

    def _append_headers(self, sheet):
        sheet.append([
            "Name", "Category", "Distance", "Position", "Time", "Date", "Race",
            "Location", "Overall Position", "Age Position", "Gender Position"
        ])

        return sheet

    def _append_new_rows(self, curr_results, prev_results, sheet):
        last_row_added = None
        for row in curr_results.iter_rows(values_only=True, min_row=2):
            row = self._normalize_row(row)
            if row[5] == "parkrun":
                continue

            seen = False
            for old_row in prev_results.iter_rows(values_only=True, min_row=2):
                old_row = self._normalize_row(old_row)

                if row == old_row:
                    seen = True
                    break

            if not seen:
                # If the event is new, insert a new section
                if last_row_added and self._is_new_event(row, last_row_added):
                    sheet.append([])  # Add a blank row
                    sheet = self._append_headers(sheet)

                sheet.append(self._format_row(row))
                last_row_added = row

        return sheet

    def _is_new_event(self, curr_row, prev_row):
        # Check if the event name, location and date are different
        return (curr_row[4], curr_row[6:7]) != (prev_row[4], prev_row[6:7])

    def _normalize_row(self, row):
        # Remove whitespace and convert empty cells to None

        row = tuple(c.strip() if isinstance(c, str) else c for c in row)
        return tuple(None if c == "" else c for c in row)

    def _format_row(self, row):
        [
            first_name,
            surname,
            sex,
            age_category,
            date,
            distance,
            race,
            location,
            overall_position,
            age_position,
            gender_postion,
            _best,
            _age_grading,
            _club_record,
            time,
        ] = row

        position = gender_postion if gender_postion else overall_position
        return [
            f"{first_name} {surname}", f"{sex}{age_category}", distance,
            position, time, date, race, location, overall_position,
            age_position, gender_postion
        ]

    def save(self, filename):
        if self.results is not None:
            file = File.objects.create(file=filename)
            self.results.save(file.file.path)
        else:
            raise Exception("No results to save")
