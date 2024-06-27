from unittest.mock import patch
from xmlrpc.client import Boolean

from django.test import TestCase
from faker import Faker
from files.models import File
from openpyxl import Workbook
from results.results_processor import ResultsProcessor

fake = Faker()


def fake_results(num_results: int, seed: int = 1981) -> Workbook:
    fake.seed_instance(seed)

    wb = Workbook()
    wb.remove(wb[wb.sheetnames[0]])
    sheet = wb.create_sheet("Results (30 days)")

    sheet.append([
        "First Name",
        "Surname",
        "Sex",
        "Age Category",
        "Date",
        "Distance",
        "Race",
        "Location",
        "Position",
        "Age Position",
        "Gender Position",
        "Best",
        "Age Grading",
        "Club Record",
        "Time",
    ])

    for _ in range(num_results):
        sheet.append(fake_result())

    return wb


def fake_result(parkrun: Boolean = False) -> tuple:
    return (
        fake.first_name(),
        fake.last_name(),
        fake.random_element(["M", "F"]),
        fake.random_element(["U13", "SEN", "V35", "V50", "V60"]),
        fake.date(),
        "parkrun"
        if parkrun else fake.random_element(["5K", "10K", "HM", "M"]),
        fake.word(),
        fake.city(),
        fake.random_int(),
        fake.random_int(),
        fake.random_int(),
        fake.random_element(["PB", None]),
        fake.random_int(0, 100),
        fake.random_element(["Yes", None]),
        fake.time(),
    )


class TestProcessResults(TestCase):

    def test_fetch_results_returns_an_empty_list_when_no_results(self):
        self.assertEqual(0, len(ResultsProcessor.fetch_results(2)))

    def test_fetch_results_returns_expected_number_of_files(self):
        File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        File.objects.create(file="BrightonPhoenix_2024-05-12.xlsx")
        self.assertEqual(2, len(ResultsProcessor.fetch_results(2)))

    def test_fetch_results_returns_most_recent_results(self):
        File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        second_file = File.objects.create(
            file="BrightonPhoenix_2024-05-12.xlsx")
        self.assertEqual([second_file], ResultsProcessor.fetch_results(1))

    def test_fetch_results_ignores_non_results_files(self):
        File.objects.create(file="not-results.jpg")
        self.assertListEqual([], ResultsProcessor.fetch_results(1))

    def test_process_creates_new_sheet_with_new_results(self):
        old_file = File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        new_file = File.objects.create(file="BrightonPhoenix_2024-05-12.xlsx")

        with patch("results.results_processor.load_workbook") as mock:
            mock.side_effect = [fake_results(15), fake_results(10)]

            processor = ResultsProcessor()
            processor.process(new_file, old_file)

            assert processor.results is not None
            self.assertEqual(2, len(processor.results.sheetnames))
            self.assertEqual("NEW RESULTS", processor.results.sheetnames[0])
            self.assertEqual(6,
                             len(list(processor.results["NEW RESULTS"].rows)))

    def test_process_populates_new_sheet_with_expected_results(self):
        old_file = File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        new_file = File.objects.create(file="BrightonPhoenix_2024-05-12.xlsx")

        with patch("results.results_processor.load_workbook") as mock:
            prev, curr = (fake_results(10), fake_results(10))
            [
                first_name, surname, sex, age_category, date, distance, race,
                location, _, _, position, _, _, _, time
            ] = new_result = fake_result()
            curr["Results (30 days)"].append(new_result)

            mock.side_effect = [curr, prev]

            processor = ResultsProcessor()
            processor.process(new_file, old_file)

            assert processor.results is not None
            sheet = processor.results["NEW RESULTS"]
            new_results = list(sheet.rows)
            headers = list(cell.value for cell in new_results[0])
            expected_headers = [
                "Name",
                "Category",
                "Distance",
                "Position",
                "Time",
                "Date",
                "Race",
                "Location",
            ]

            self.assertEqual(headers, expected_headers)

            result = list(cell.value for cell in new_results[1])
            expected_result = [
                f"{first_name} {surname}",
                f"{sex}{age_category}",
                distance,
                position,
                time,
                date,
                race,
                location,
            ]

            self.assertEqual(result, expected_result)

    def test_process_ignores_parkruns(self):
        old_file = File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        new_file = File.objects.create(file="BrightonPhoenix_2024-05-12.xlsx")

        with patch("results.results_processor.load_workbook") as mock:
            prev, curr = (fake_results(10), fake_results(10))
            new_result = fake_result(parkrun=True)
            curr["Results (30 days)"].append(new_result)

            mock.side_effect = [curr, prev]

            processor = ResultsProcessor()
            processor.process(new_file, old_file)

            assert processor.results is not None
            sheet = processor.results["NEW RESULTS"]
            self.assertEqual(1, len(list(sheet.rows)))

    def test_process_ignores_whitespace(self):
        old_file = File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        new_file = File.objects.create(file="BrightonPhoenix_2024-05-12.xlsx")

        with patch("results.results_processor.load_workbook") as mock:
            prev, curr = (fake_results(10), fake_results(10))
            res = fake_result()
            leading_whitespace = tuple(f" {value}" for value in res)
            trailing_whitespace = tuple(f"{value} " for value in res)

            prev["Results (30 days)"].append(leading_whitespace)
            curr["Results (30 days)"].append(trailing_whitespace)

            mock.side_effect = [curr, prev]

            processor = ResultsProcessor()
            processor.process(new_file, old_file)

            assert processor.results is not None
            sheet = processor.results["NEW RESULTS"]

            # only the header row, no new results
            self.assertEqual(1, len(list(sheet.rows)))

    def test_save_creates_new_file(self):
        old_file = File.objects.create(file="BrightonPhoenix_2024-05-05.xlsx")
        new_file = File.objects.create(file="BrightonPhoenix_2024-05-12.xlsx")

        with patch("results.results_processor.load_workbook") as mock:
            mock.side_effect = [fake_results(10), fake_results(10)]

            processor = ResultsProcessor()
            processor.process(new_file, old_file)

            filename = "new_results.xlsx"
            with patch("openpyxl.workbook.Workbook.save") as mock_save:
                processor.save(filename)
                mock_save.assert_called_once()
                self.assertTrue(mock_save.call_args[0][0].endswith(filename))

            self.assertIsNotNone(File.objects.get(file=filename))

    def test_save_raises_exception_when_no_results_to_save(self):
        processor = ResultsProcessor()
        with self.assertRaises(Exception):
            processor.save("new_results.xlsx")
